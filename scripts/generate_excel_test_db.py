"""
Excel로 테스트 DB 생성 스크립트

법인 3개, 직원 60명(각 법인당 20명, 남녀 각 10명), 개인 20명(남녀 각 10명)의 
완전한 테스트 데이터를 Excel 파일로 생성합니다.

각 테이블을 하나의 시트로 구성하여 관계를 명확히 합니다.

실행 방법:
    python scripts/generate_excel_test_db.py
"""
import sys
import os
import random
import pandas as pd
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 샘플 파일 경로
SAMPLE_BASE_PATH = r"D:\projects\_operation_docs\dev_docs_hr\sample"
RRN_EXCEL_PATH = os.path.join(SAMPLE_BASE_PATH, "주민번호.xlsx")

# 출력 파일 경로
OUTPUT_EXCEL_PATH = r"D:\projects\hrmanagement\test_db_with_hr.xlsx"

# ============================================================
# 샘플 데이터 상수
# ============================================================

COMPANY_NAMES = [
    '주식회사 테크노비전',
    '주식회사 글로벌솔루션',
    '주식회사 디지털인사이트'
]

MALE_NAMES = [
    '김민수', '이준호', '박성우', '최동현', '정태영', '강지훈', '윤상우', '장현우', '임재현', '한지훈',
    '김도현', '이승민', '박준혁', '최민석', '정우진', '강태현', '윤지훈', '장민규', '임성민', '한동욱',
    '김현우', '이진호', '박건우', '최성민', '정민재', '강태영', '윤준서', '장현석', '임동혁', '한준영'
]

FEMALE_NAMES = [
    '김지은', '이수진', '박민지', '최예진', '정소영', '강지원', '윤서연', '장하늘', '임지현', '한수빈',
    '김다은', '이채원', '박서윤', '최지우', '정유진', '강민서', '윤지원', '장예나', '임소은', '한나은',
    '김하늘', '이서영', '박지민', '최은지', '정수아', '강예린', '윤채린', '장소연', '임지안', '한서아'
]

DEPARTMENTS = ['개발팀', '기획팀', '디자인팀', '마케팅팀', '영업팀', '인사팀', '재무팀', 'R&D팀']
POSITIONS = ['인턴', '사원', '주임', '대리', '과장', '차장', '부장', '이사']
JOB_TITLES = ['팀원', '파트장', '팀장', '본부장', '이사', '대표']
STATUSES = ['재직', '휴직', '퇴사']
GENDERS = ['남', '여']

EDUCATION_LEVELS = ['고등학교', '전문대학', '대학교', '대학원(석사)', '대학원(박사)']
UNIVERSITIES = [
    '서울대학교', '연세대학교', '고려대학교', '한국과학기술원', '포스텍',
    '성균관대학교', '한양대학교', '중앙대학교', '경희대학교', '이화여자대학교'
]
MAJORS = [
    '컴퓨터공학', '전자공학', '산업공학', '경영학', '경제학',
    '심리학', '디자인', '마케팅', '회계학', '법학'
]

CAREER_COMPANIES = [
    '네이버', '카카오', '삼성전자', 'LG전자', 'SK하이닉스',
    '현대자동차', '기아자동차', '롯데', '신세계', 'CJ'
]

CERTIFICATE_NAMES = [
    '정보처리기사', '정보처리산업기사', 'SQLD', 'SQLP', 'AWS 자격증',
    'PMP', '컴퓨터활용능력', '토익', '토플', '오픽'
]

LANGUAGES = ['영어', '일본어', '중국어', '스페인어', '프랑스어', '독일어']
LANGUAGE_LEVELS = ['초급', '중급', '고급', '원어민수준']

MILITARY_STATUS = ['군필', '미필', '면제', '해당없음']
SERVICE_TYPES = ['육군', '해군', '공군', '해병대', '의무경찰', '사회복무요원', '산업기능요원']
RANKS = ['이병', '일병', '상병', '병장', '하사', '중사']

PROMOTION_TYPES = ['승진', '전보', '발령', '파견', '복직', '휴직', '직무변경']
EVALUATION_GRADES = ['S', 'A', 'B', 'C', 'D']

TRAINING_NAMES = [
    '신입사원 교육', '리더십 교육', 'OJT 교육', '직무역량 교육', '안전교육',
    '성희롱 예방교육', '개인정보보호 교육', '정보보안 교육', 'AI/ML 기초',
    '프로젝트 관리', '커뮤니케이션 스킬', '프레젠테이션 기법', '데이터 분석 실무'
]

TRAINING_INSTITUTIONS = [
    '사내교육팀', '한국생산성본부', '한국능률협회', '삼성멀티캠퍼스',
    'Coursera', 'Udemy', '패스트캠퍼스', '인프런'
]

PROJECT_NAMES = [
    'ERP 시스템 구축', '모바일 앱 개발', '웹사이트 리뉴얼', 'AI 챗봇 개발',
    '클라우드 마이그레이션', '보안 시스템 강화', '데이터 분석 플랫폼',
    'CRM 시스템 도입', 'IoT 플랫폼 개발', 'DevOps 환경 구축',
    'HR 시스템 고도화', '전자결재 시스템 구축', '통합 로그 시스템'
]

PROJECT_ROLES = ['PM', 'PL', '개발자', '기획자', 'QA', '디자이너', 'DBA', '아키텍트']

ASSET_ITEMS = [
    ('노트북', ['MacBook Pro 14', 'ThinkPad X1 Carbon', 'Dell XPS 15', 'LG Gram 17']),
    ('모니터', ['LG 27UL850', 'Dell U2720Q', 'Samsung 32 Curved']),
    ('키보드', ['Apple Magic Keyboard', 'Logitech MX Keys', 'HHKB Professional']),
    ('마우스', ['Apple Magic Mouse', 'Logitech MX Master 3', 'Microsoft Arc Mouse']),
    ('휴대폰', ['iPhone 15 Pro', 'Galaxy S24 Ultra', 'Pixel 8 Pro']),
    ('의자', ['허먼밀러 에어론', '시디즈 T80', '듀오백 DK-2500G']),
]

AWARD_NAMES = [
    '이달의 우수사원', '분기 MVP', '프로젝트 공로상', '고객만족상',
    '혁신상', '봉사상', '장기근속상', '신인상', '팀워크상'
]

FAMILY_RELATIONS = ['배우자', '자녀', '부', '모', '형', '누나', '동생', '조부', '조모']

# ============================================================
# 유틸리티 함수
# ============================================================

def load_resident_numbers():
    """주민번호.xlsx에서 주민번호 리스트 로드"""
    try:
        if not os.path.exists(RRN_EXCEL_PATH):
            print(f"  [WARNING] 주민번호 Excel 파일 없음: {RRN_EXCEL_PATH}")
            return []
        df = pd.read_excel(RRN_EXCEL_PATH)
        # 첫 번째 컬럼을 주민번호로 가정
        rrn_list = df.iloc[:, 0].dropna().astype(str).tolist()
        return rrn_list
    except Exception as e:
        print(f"  [ERROR] 주민번호 로드 실패: {e}")
        return []


def generate_phone():
    """랜덤 휴대폰 번호 생성"""
    return f"010-{random.randint(1000, 9999)}-{random.randint(1000, 9999)}"


def generate_email(name, domain='company.com'):
    """이메일 생성"""
    import unicodedata
    romanized = ''.join(
        c for c in unicodedata.normalize('NFD', name)
        if unicodedata.category(c) != 'Mn'
    ).replace(' ', '').lower()
    if not romanized or not romanized.isascii():
        romanized = f"user{random.randint(1000, 9999)}"
    return f"{romanized}@{domain}"


def generate_birth_date(min_age=25, max_age=55):
    """생년월일 생성"""
    today = date.today()
    age = random.randint(min_age, max_age)
    birth_year = today.year - age
    birth_month = random.randint(1, 12)
    birth_day = random.randint(1, 28)
    return f"{birth_year}-{birth_month:02d}-{birth_day:02d}"


def generate_hire_date(min_years=0, max_years=15):
    """입사일 생성"""
    today = date.today()
    years_ago = random.randint(min_years, max_years)
    months_ago = random.randint(0, 11)
    hire_date = today - timedelta(days=(years_ago * 365 + months_ago * 30))
    return hire_date.strftime('%Y-%m-%d')


def generate_past_date(min_years=1, max_years=5):
    """과거 날짜 생성"""
    today = date.today()
    days_ago = random.randint(min_years * 365, max_years * 365)
    past_date = today - timedelta(days=days_ago)
    return past_date.strftime('%Y-%m-%d')


def generate_address():
    """주소 생성"""
    cities = ['서울특별시', '부산광역시', '대구광역시', '인천광역시', '경기도']
    districts = {
        '서울특별시': ['강남구', '서초구', '송파구', '마포구', '영등포구'],
        '부산광역시': ['해운대구', '수영구', '동래구', '부산진구'],
        '대구광역시': ['수성구', '달서구', '중구'],
        '인천광역시': ['연수구', '남동구', '부평구'],
        '경기도': ['성남시 분당구', '용인시 수지구', '수원시 영통구'],
    }
    city = random.choice(cities)
    district = random.choice(districts.get(city, ['중구']))
    return f"{city} {district} 테스트로 {random.randint(1, 500)}"


def generate_postal_code():
    """우편번호 생성"""
    return f"{random.randint(10000, 99999)}"


def generate_employee_number(index):
    """사번 생성"""
    year = datetime.now().year
    return f"EMP-{year}-{index:04d}"


def generate_business_number():
    """사업자등록번호 생성"""
    return f"{random.randint(100, 999)}{random.randint(10, 99)}{random.randint(10000, 99999)}"


# ============================================================
# 데이터 생성 함수
# ============================================================

def create_companies():
    """법인 데이터 생성"""
    companies = []
    business_numbers = ['1234567890', '2345678901', '3456789012']
    
    for idx, name in enumerate(COMPANY_NAMES):
        companies.append({
            'id': idx + 1,
            'name': name,
            'business_number': business_numbers[idx],
            'establishment_date': f"{random.randint(2000, 2015)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}",
            'address': generate_address(),
            'phone': generate_phone(),
            'email': f"contact@{name.lower().replace('주식회사 ', '').replace(' ', '')}.com",
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return companies


def create_organizations(companies):
    """조직 데이터 생성"""
    organizations = []
    org_id = 1
    
    for company in companies:
        # 루트 조직
        organizations.append({
            'id': org_id,
            'company_id': company['id'],
            'name': company['name'],
            'code': f"ORG-{company['id']:03d}-ROOT",
            'parent_id': None,
            'level': 0,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        root_id = org_id
        org_id += 1
        
        # 부서 조직 (각 법인당 3-5개)
        num_depts = random.randint(3, 5)
        for dept in random.sample(DEPARTMENTS, num_depts):
            organizations.append({
                'id': org_id,
                'company_id': company['id'],
                'name': dept,
                'code': f"ORG-{company['id']:03d}-{org_id:03d}",
                'parent_id': root_id,
                'level': 1,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            org_id += 1
    
    return organizations


def create_users_and_employees(companies, organizations, resident_numbers):
    """사용자 및 직원 데이터 생성"""
    users = []
    employees = []
    profiles = []
    educations = []
    careers = []
    certificates = []
    languages = []
    military_services = []
    # 인사관리 테이블
    contracts = []
    promotions = []
    evaluations = []
    trainings = []
    attendances = []
    salaries = []
    salary_histories = []
    salary_payments = []
    benefits = []
    insurances = []
    assets = []
    awards = []
    family_members = []
    hr_projects = []
    project_participations = []
    
    user_id = 1
    emp_id = 1
    profile_id = 1
    edu_id = 1
    career_id = 1
    cert_id = 1
    lang_id = 1
    mil_id = 1
    # 인사관리 ID 카운터
    contract_id = 1
    promo_id = 1
    eval_id = 1
    training_id = 1
    attendance_id = 1
    salary_id = 1
    salary_history_id = 1
    salary_payment_id = 1
    benefit_id = 1
    insurance_id = 1
    asset_id = 1
    award_id = 1
    family_id = 1
    hr_project_id = 1
    project_participation_id = 1
    
    # 법인별 조직 리스트
    company_orgs = {}
    for org in organizations:
        if org['parent_id'] is not None:  # 루트가 아닌 조직만
            if org['company_id'] not in company_orgs:
                company_orgs[org['company_id']] = []
            company_orgs[org['company_id']].append(org)
    
    rrn_idx = 0
    
    for company_idx, company in enumerate(companies):
        # 남성 직원 10명
        for male_idx in range(10):
            name = MALE_NAMES[company_idx * 10 + male_idx]
            gender = '남'
            birth_date = generate_birth_date()
            
            # 주민번호
            if rrn_idx < len(resident_numbers):
                rrn = resident_numbers[rrn_idx]
                rrn_idx += 1
            else:
                rrn = f"{birth_date.replace('-', '')[2:]}-1{random.randint(100000, 999999)}"
            
            # 조직 할당
            org_list = company_orgs.get(company['id'], [])
            org = random.choice(org_list) if org_list else None
            
            # User 생성
            username = f"emp_{company_idx + 1}_{male_idx + 1}"
            users.append({
                'id': user_id,
                'username': username,
                'email': generate_email(name, f"company{company_idx + 1}.com"),
                'password_hash': 'pbkdf2:sha256:600000$dummy$dummy',
                'role': 'employee',
                'account_type': 'employee_sub',
                'company_id': company['id'],
                'employee_id': emp_id,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Profile 생성
            profiles.append({
                'id': profile_id,
                'name': name,
                'gender': gender,
                'birth_date': birth_date,
                'resident_number': rrn,
                'mobile_phone': generate_phone(),
                'email': generate_email(name),
                'address': generate_address(),
                'postal_code': generate_postal_code(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Employee 생성
            employees.append({
                'id': emp_id,
                'employee_number': generate_employee_number(emp_id),
                'name': name,
                'profile_id': profile_id,
                'company_id': company['id'],
                'organization_id': org['id'] if org else None,
                'department': org['name'] if org else random.choice(DEPARTMENTS),
                'position': random.choice(POSITIONS),
                'job_title': random.choice(JOB_TITLES),
                'status': '재직',
                'hire_date': generate_hire_date(),
                'phone': generate_phone(),
                'email': generate_email(name, f"company{company_idx + 1}.com"),
                'gender': gender,
                'birth_date': birth_date,
                'mobile_phone': generate_phone(),
                'address': generate_address(),
                'postal_code': generate_postal_code(),
                'resident_number': rrn,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Education (1-2개)
            for _ in range(random.randint(1, 2)):
                educations.append({
                    'id': edu_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'level': random.choice(EDUCATION_LEVELS),
                    'school_name': random.choice(UNIVERSITIES),
                    'major': random.choice(MAJORS),
                    'graduation_date': f"{random.randint(2010, 2020)}-{random.randint(1, 12):02d}",
                    'status': '졸업',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                edu_id += 1
            
            # Career (1-3개)
            for _ in range(random.randint(1, 3)):
                start_date = generate_hire_date(max_years=10)
                end_date = generate_hire_date(min_years=0, max_years=5)
                careers.append({
                    'id': career_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'company_name': random.choice(CAREER_COMPANIES),
                    'position': random.choice(POSITIONS),
                    'start_date': start_date,
                    'end_date': end_date,
                    'description': '담당 업무 수행',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                career_id += 1
            
            # Certificate (0-2개)
            for _ in range(random.randint(0, 2)):
                cert_date = generate_hire_date(max_years=5)
                certificates.append({
                    'id': cert_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'name': random.choice(CERTIFICATE_NAMES),
                    'issuing_organization': '한국산업인력공단',
                    'issue_date': cert_date,
                    'expiry_date': None,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                cert_id += 1
            
            # Language (0-2개)
            for _ in range(random.randint(0, 2)):
                languages.append({
                    'id': lang_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'language': random.choice(LANGUAGES),
                    'level': random.choice(LANGUAGE_LEVELS),
                    'test_name': 'TOEIC',
                    'test_score': random.randint(600, 990) if random.choice(LANGUAGES) == '영어' else None,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                lang_id += 1
            
            # Military Service
            if gender == '남':
                military_services.append({
                    'id': mil_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'status': random.choice(['군필', '면제']),
                    'service_type': random.choice(SERVICE_TYPES) if random.random() > 0.3 else None,
                    'rank': random.choice(RANKS) if random.random() > 0.3 else None,
                    'start_date': f"{random.randint(2010, 2018)}-{random.randint(1, 12):02d}",
                    'end_date': f"{random.randint(2012, 2020)}-{random.randint(1, 12):02d}",
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                mil_id += 1
            
            # 인사관리 데이터 생성
            hire_date = employees[-1]['hire_date']
            current_year = datetime.now().year
            
            # Contract (계약)
            contracts.append({
                'id': contract_id,
                'employee_id': emp_id,
                'contract_type': '정규직',
                'contract_date': hire_date,
                'contract_period': '무기계약',
                'employee_type': '정규직',
                'work_type': '주5일',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            contract_id += 1
            
            # Promotion (발령/인사이동) - 1-3개
            current_dept = employees[-1]['department']
            current_pos = employees[-1]['position']
            for _ in range(random.randint(1, 3)):
                new_dept = random.choice(DEPARTMENTS)
                new_pos = random.choice(POSITIONS)
                promo_date = generate_hire_date(max_years=5)
                promotions.append({
                    'id': promo_id,
                    'employee_id': emp_id,
                    'effective_date': promo_date,
                    'promotion_type': random.choice(PROMOTION_TYPES),
                    'from_department': current_dept,
                    'to_department': new_dept,
                    'from_position': current_pos,
                    'to_position': new_pos,
                    'job_role': random.choice(JOB_TITLES),
                    'reason': '정기 인사',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                promo_id += 1
                current_dept, current_pos = new_dept, new_pos
            
            # Evaluation (인사평가) - 최근 3년
            for year in range(current_year - 2, current_year + 1):
                evaluations.append({
                    'id': eval_id,
                    'employee_id': emp_id,
                    'year': year,
                    'q1_grade': random.choice(EVALUATION_GRADES),
                    'q2_grade': random.choice(EVALUATION_GRADES),
                    'q3_grade': random.choice(EVALUATION_GRADES) if year < current_year else None,
                    'q4_grade': random.choice(EVALUATION_GRADES) if year < current_year else None,
                    'overall_grade': random.choice(EVALUATION_GRADES) if year < current_year else None,
                    'salary_negotiation': f"+{random.randint(0, 10)}%" if year < current_year else None,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                eval_id += 1
            
            # Training (교육 이력) - 2-5개
            for _ in range(random.randint(2, 5)):
                trainings.append({
                    'id': training_id,
                    'employee_id': emp_id,
                    'training_date': generate_hire_date(max_years=3),
                    'training_name': random.choice(TRAINING_NAMES),
                    'institution': random.choice(TRAINING_INSTITUTIONS),
                    'hours': random.choice([8, 16, 24, 40, 80]),
                    'completion_status': random.choice(['수료', '이수', '진행중']),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                training_id += 1
            
            # Attendance (근태) - 최근 12개월
            for month_offset in range(12):
                month_date = date.today().replace(day=1) - timedelta(days=30 * month_offset)
                attendances.append({
                    'id': attendance_id,
                    'employee_id': emp_id,
                    'year': month_date.year,
                    'month': month_date.month,
                    'work_days': random.randint(18, 23),
                    'absent_days': random.randint(0, 2),
                    'late_count': random.randint(0, 3),
                    'early_leave_count': random.randint(0, 2),
                    'annual_leave_used': random.randint(0, 2),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                attendance_id += 1
            
            # Salary (급여 정보)
            base_salary = random.randint(3000, 8000) * 10000
            position_allowance = random.randint(10, 50) * 10000
            meal_allowance = 200000
            transportation_allowance = 100000
            total_salary = base_salary + position_allowance + meal_allowance + transportation_allowance
            
            salaries.append({
                'id': salary_id,
                'employee_id': emp_id,
                'salary_type': random.choice(['연봉제', '월급제']),
                'base_salary': base_salary,
                'position_allowance': position_allowance,
                'meal_allowance': meal_allowance,
                'transportation_allowance': transportation_allowance,
                'total_salary': total_salary,
                'payment_day': 25,
                'payment_method': '계좌이체',
                'annual_salary': total_salary * 12,
                'monthly_salary': total_salary,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            salary_id += 1
            
            # SalaryHistory (연봉 이력) - 최근 3년
            for year in range(current_year - 2, current_year + 1):
                annual = base_salary * 12 * (1 + (current_year - year) * 0.05)
                salary_histories.append({
                    'id': salary_history_id,
                    'employee_id': emp_id,
                    'contract_year': year,
                    'annual_salary': int(annual),
                    'bonus': int(annual * 0.1),
                    'total_amount': int(annual * 1.1),
                    'contract_period': f"{year}-01-01 ~ {year}-12-31",
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                salary_history_id += 1
            
            # SalaryPayment (급여 지급 이력) - 최근 6개월
            for month_offset in range(6):
                payment_date = date.today().replace(day=25) - timedelta(days=30 * month_offset)
                gross = total_salary
                insurance_amount = int(gross * 0.09)
                tax = int(gross * 0.03)
                deduction = insurance_amount + tax
                salary_payments.append({
                    'id': salary_payment_id,
                    'employee_id': emp_id,
                    'payment_date': payment_date.strftime('%Y-%m-%d'),
                    'payment_period': f"{payment_date.year}년 {payment_date.month}월",
                    'base_salary': base_salary,
                    'allowances': position_allowance + meal_allowance + transportation_allowance,
                    'gross_pay': gross,
                    'insurance': insurance_amount,
                    'income_tax': tax,
                    'total_deduction': deduction,
                    'net_pay': gross - deduction,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                salary_payment_id += 1
            
            # Insurance (보험 정보)
            insurances.append({
                'id': insurance_id,
                'employee_id': emp_id,
                'national_pension': True,
                'health_insurance': True,
                'employment_insurance': True,
                'industrial_accident': True,
                'national_pension_rate': 4.5,
                'health_insurance_rate': 3.545,
                'long_term_care_rate': 0.9182,
                'employment_insurance_rate': 0.9,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            insurance_id += 1
            
            # Benefit (복리후생)
            hire_years = (date.today() - datetime.strptime(hire_date, '%Y-%m-%d').date()).days // 365
            annual_leave = min(15 + hire_years, 25)
            benefits.append({
                'id': benefit_id,
                'employee_id': emp_id,
                'year': current_year,
                'annual_leave_granted': annual_leave,
                'annual_leave_used': random.randint(0, annual_leave // 2),
                'annual_leave_remaining': annual_leave - random.randint(0, annual_leave // 2),
                'severance_type': '퇴직연금',
                'severance_method': 'DB형',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            benefit_id += 1
            
            # Asset (자산 배정) - 1-3개
            for _ in range(random.randint(1, 3)):
                item_category, models = random.choice(ASSET_ITEMS)
                assets.append({
                    'id': asset_id,
                    'employee_id': emp_id,
                    'issue_date': generate_hire_date(max_years=2),
                    'item_name': item_category,
                    'model': random.choice(models),
                    'serial_number': f"SN{random.randint(100000, 999999)}",
                    'status': random.choice(['사용중', '반납', '수리중']),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                asset_id += 1
            
            # Award (수상) - 0-2개
            for _ in range(random.randint(0, 2)):
                awards.append({
                    'id': award_id,
                    'employee_id': emp_id,
                    'award_date': generate_hire_date(max_years=3),
                    'award_name': random.choice(AWARD_NAMES),
                    'institution': company['name'],
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                award_id += 1
            
            # FamilyMember (가족) - 0-4명
            if random.random() > 0.3:  # 70% 확률
                for _ in range(random.randint(1, 4)):
                    family_members.append({
                        'id': family_id,
                        'employee_id': emp_id,
                        'relation': random.choice(FAMILY_RELATIONS),
                        'name': f"가족{random.randint(1, 10)}",
                        'birth_date': generate_birth_date(1, 80),
                        'occupation': random.choice(['회사원', '공무원', '자영업', '학생', '주부', None]),
                        'contact': generate_phone() if random.random() > 0.5 else None,
                        'is_cohabitant': random.choice([True, False]),
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    family_id += 1
            
            # HrProject (인사이력 프로젝트) - 1-3개
            for _ in range(random.randint(1, 3)):
                start_date = generate_hire_date(max_years=2)
                hr_projects.append({
                    'id': hr_project_id,
                    'employee_id': emp_id,
                    'project_name': random.choice(PROJECT_NAMES),
                    'start_date': start_date,
                    'end_date': None if random.random() > 0.7 else generate_hire_date(max_years=1),
                    'duration': f"{random.randint(3, 24)}개월",
                    'role': random.choice(PROJECT_ROLES),
                    'duty': f"{random.choice(['개발', '기획', '관리', '분석'])} 업무",
                    'client': f"{random.choice(['삼성', 'LG', 'SK', 'KT', '현대', '롯데'])}그룹",
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                hr_project_id += 1
            
            # ProjectParticipation (경력 프로젝트 참여) - 0-2개
            for _ in range(random.randint(0, 2)):
                project_participations.append({
                    'id': project_participation_id,
                    'employee_id': emp_id,
                    'project_name': f"이전회사 {random.choice(PROJECT_NAMES)}",
                    'start_date': generate_hire_date(max_years=5),
                    'end_date': generate_hire_date(max_years=3),
                    'role': random.choice(PROJECT_ROLES),
                    'duty': f"{random.choice(['개발', '기획', '관리'])} 담당",
                    'client': '이전고객사',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                project_participation_id += 1
            
            user_id += 1
            emp_id += 1
            profile_id += 1
        
        # 여성 직원 10명
        for female_idx in range(10):
            name = FEMALE_NAMES[company_idx * 10 + female_idx]
            gender = '여'
            birth_date = generate_birth_date()
            
            # 주민번호
            if rrn_idx < len(resident_numbers):
                rrn = resident_numbers[rrn_idx]
                rrn_idx += 1
            else:
                rrn = f"{birth_date.replace('-', '')[2:]}-2{random.randint(100000, 999999)}"
            
            # 조직 할당
            org_list = company_orgs.get(company['id'], [])
            org = random.choice(org_list) if org_list else None
            
            # User 생성
            username = f"emp_{company_idx + 1}_{female_idx + 11}"
            users.append({
                'id': user_id,
                'username': username,
                'email': generate_email(name, f"company{company_idx + 1}.com"),
                'password_hash': 'pbkdf2:sha256:600000$dummy$dummy',
                'role': 'employee',
                'account_type': 'employee_sub',
                'company_id': company['id'],
                'employee_id': emp_id,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Profile 생성
            profiles.append({
                'id': profile_id,
                'name': name,
                'gender': gender,
                'birth_date': birth_date,
                'resident_number': rrn,
                'mobile_phone': generate_phone(),
                'email': generate_email(name),
                'address': generate_address(),
                'postal_code': generate_postal_code(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Employee 생성
            employees.append({
                'id': emp_id,
                'employee_number': generate_employee_number(emp_id),
                'name': name,
                'profile_id': profile_id,
                'company_id': company['id'],
                'organization_id': org['id'] if org else None,
                'department': org['name'] if org else random.choice(DEPARTMENTS),
                'position': random.choice(POSITIONS),
                'job_title': random.choice(JOB_TITLES),
                'status': '재직',
                'hire_date': generate_hire_date(),
                'phone': generate_phone(),
                'email': generate_email(name, f"company{company_idx + 1}.com"),
                'gender': gender,
                'birth_date': birth_date,
                'mobile_phone': generate_phone(),
                'address': generate_address(),
                'postal_code': generate_postal_code(),
                'resident_number': rrn,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Education (1-2개)
            for _ in range(random.randint(1, 2)):
                educations.append({
                    'id': edu_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'level': random.choice(EDUCATION_LEVELS),
                    'school_name': random.choice(UNIVERSITIES),
                    'major': random.choice(MAJORS),
                    'graduation_date': f"{random.randint(2010, 2020)}-{random.randint(1, 12):02d}",
                    'status': '졸업',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                edu_id += 1
            
            # Career (1-3개)
            for _ in range(random.randint(1, 3)):
                start_date = generate_hire_date(max_years=10)
                end_date = generate_hire_date(min_years=0, max_years=5)
                careers.append({
                    'id': career_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'company_name': random.choice(CAREER_COMPANIES),
                    'position': random.choice(POSITIONS),
                    'start_date': start_date,
                    'end_date': end_date,
                    'description': '담당 업무 수행',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                career_id += 1
            
            # Certificate (0-2개)
            for _ in range(random.randint(0, 2)):
                cert_date = generate_hire_date(max_years=5)
                certificates.append({
                    'id': cert_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'name': random.choice(CERTIFICATE_NAMES),
                    'issuing_organization': '한국산업인력공단',
                    'issue_date': cert_date,
                    'expiry_date': None,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                cert_id += 1
            
            # Language (0-2개)
            for _ in range(random.randint(0, 2)):
                languages.append({
                    'id': lang_id,
                    'employee_id': emp_id,
                    'profile_id': profile_id,
                    'language': random.choice(LANGUAGES),
                    'level': random.choice(LANGUAGE_LEVELS),
                    'test_name': 'TOEIC',
                    'test_score': random.randint(600, 990) if random.choice(LANGUAGES) == '영어' else None,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                lang_id += 1
            
            # 인사관리 데이터 생성 (여성 직원)
            hire_date = employees[-1]['hire_date']
            current_year = datetime.now().year
            
            # Contract (계약)
            contracts.append({
                'id': contract_id,
                'employee_id': emp_id,
                'contract_type': '정규직',
                'contract_date': hire_date,
                'contract_period': '무기계약',
                'employee_type': '정규직',
                'work_type': '주5일',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            contract_id += 1
            
            # Promotion (발령/인사이동) - 1-3개
            current_dept = employees[-1]['department']
            current_pos = employees[-1]['position']
            for _ in range(random.randint(1, 3)):
                new_dept = random.choice(DEPARTMENTS)
                new_pos = random.choice(POSITIONS)
                promo_date = generate_hire_date(max_years=5)
                promotions.append({
                    'id': promo_id,
                    'employee_id': emp_id,
                    'effective_date': promo_date,
                    'promotion_type': random.choice(PROMOTION_TYPES),
                    'from_department': current_dept,
                    'to_department': new_dept,
                    'from_position': current_pos,
                    'to_position': new_pos,
                    'job_role': random.choice(JOB_TITLES),
                    'reason': '정기 인사',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                promo_id += 1
                current_dept, current_pos = new_dept, new_pos
            
            # Evaluation (인사평가) - 최근 3년
            for year in range(current_year - 2, current_year + 1):
                evaluations.append({
                    'id': eval_id,
                    'employee_id': emp_id,
                    'year': year,
                    'q1_grade': random.choice(EVALUATION_GRADES),
                    'q2_grade': random.choice(EVALUATION_GRADES),
                    'q3_grade': random.choice(EVALUATION_GRADES) if year < current_year else None,
                    'q4_grade': random.choice(EVALUATION_GRADES) if year < current_year else None,
                    'overall_grade': random.choice(EVALUATION_GRADES) if year < current_year else None,
                    'salary_negotiation': f"+{random.randint(0, 10)}%" if year < current_year else None,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                eval_id += 1
            
            # Training (교육 이력) - 2-5개
            for _ in range(random.randint(2, 5)):
                trainings.append({
                    'id': training_id,
                    'employee_id': emp_id,
                    'training_date': generate_hire_date(max_years=3),
                    'training_name': random.choice(TRAINING_NAMES),
                    'institution': random.choice(TRAINING_INSTITUTIONS),
                    'hours': random.choice([8, 16, 24, 40, 80]),
                    'completion_status': random.choice(['수료', '이수', '진행중']),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                training_id += 1
            
            # Attendance (근태) - 최근 12개월
            for month_offset in range(12):
                month_date = date.today().replace(day=1) - timedelta(days=30 * month_offset)
                attendances.append({
                    'id': attendance_id,
                    'employee_id': emp_id,
                    'year': month_date.year,
                    'month': month_date.month,
                    'work_days': random.randint(18, 23),
                    'absent_days': random.randint(0, 2),
                    'late_count': random.randint(0, 3),
                    'early_leave_count': random.randint(0, 2),
                    'annual_leave_used': random.randint(0, 2),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                attendance_id += 1
            
            # Salary (급여 정보)
            base_salary = random.randint(3000, 8000) * 10000
            position_allowance = random.randint(10, 50) * 10000
            meal_allowance = 200000
            transportation_allowance = 100000
            total_salary = base_salary + position_allowance + meal_allowance + transportation_allowance
            
            salaries.append({
                'id': salary_id,
                'employee_id': emp_id,
                'salary_type': random.choice(['연봉제', '월급제']),
                'base_salary': base_salary,
                'position_allowance': position_allowance,
                'meal_allowance': meal_allowance,
                'transportation_allowance': transportation_allowance,
                'total_salary': total_salary,
                'payment_day': 25,
                'payment_method': '계좌이체',
                'annual_salary': total_salary * 12,
                'monthly_salary': total_salary,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            salary_id += 1
            
            # SalaryHistory (연봉 이력) - 최근 3년
            for year in range(current_year - 2, current_year + 1):
                annual = base_salary * 12 * (1 + (current_year - year) * 0.05)
                salary_histories.append({
                    'id': salary_history_id,
                    'employee_id': emp_id,
                    'contract_year': year,
                    'annual_salary': int(annual),
                    'bonus': int(annual * 0.1),
                    'total_amount': int(annual * 1.1),
                    'contract_period': f"{year}-01-01 ~ {year}-12-31",
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                salary_history_id += 1
            
            # SalaryPayment (급여 지급 이력) - 최근 6개월
            for month_offset in range(6):
                payment_date = date.today().replace(day=25) - timedelta(days=30 * month_offset)
                gross = total_salary
                insurance_amount = int(gross * 0.09)
                tax = int(gross * 0.03)
                deduction = insurance_amount + tax
                salary_payments.append({
                    'id': salary_payment_id,
                    'employee_id': emp_id,
                    'payment_date': payment_date.strftime('%Y-%m-%d'),
                    'payment_period': f"{payment_date.year}년 {payment_date.month}월",
                    'base_salary': base_salary,
                    'allowances': position_allowance + meal_allowance + transportation_allowance,
                    'gross_pay': gross,
                    'insurance': insurance_amount,
                    'income_tax': tax,
                    'total_deduction': deduction,
                    'net_pay': gross - deduction,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                salary_payment_id += 1
            
            # Insurance (보험 정보)
            insurances.append({
                'id': insurance_id,
                'employee_id': emp_id,
                'national_pension': True,
                'health_insurance': True,
                'employment_insurance': True,
                'industrial_accident': True,
                'national_pension_rate': 4.5,
                'health_insurance_rate': 3.545,
                'long_term_care_rate': 0.9182,
                'employment_insurance_rate': 0.9,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            insurance_id += 1
            
            # Benefit (복리후생)
            hire_years = (date.today() - datetime.strptime(hire_date, '%Y-%m-%d').date()).days // 365
            annual_leave = min(15 + hire_years, 25)
            benefits.append({
                'id': benefit_id,
                'employee_id': emp_id,
                'year': current_year,
                'annual_leave_granted': annual_leave,
                'annual_leave_used': random.randint(0, annual_leave // 2),
                'annual_leave_remaining': annual_leave - random.randint(0, annual_leave // 2),
                'severance_type': '퇴직연금',
                'severance_method': 'DB형',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            benefit_id += 1
            
            # Asset (자산 배정) - 1-3개
            for _ in range(random.randint(1, 3)):
                item_category, models = random.choice(ASSET_ITEMS)
                assets.append({
                    'id': asset_id,
                    'employee_id': emp_id,
                    'issue_date': generate_hire_date(max_years=2),
                    'item_name': item_category,
                    'model': random.choice(models),
                    'serial_number': f"SN{random.randint(100000, 999999)}",
                    'status': random.choice(['사용중', '반납', '수리중']),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                asset_id += 1
            
            # Award (수상) - 0-2개
            for _ in range(random.randint(0, 2)):
                awards.append({
                    'id': award_id,
                    'employee_id': emp_id,
                    'award_date': generate_hire_date(max_years=3),
                    'award_name': random.choice(AWARD_NAMES),
                    'institution': company['name'],
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                award_id += 1
            
            # FamilyMember (가족) - 0-4명
            if random.random() > 0.3:  # 70% 확률
                for _ in range(random.randint(1, 4)):
                    family_members.append({
                        'id': family_id,
                        'employee_id': emp_id,
                        'relation': random.choice(FAMILY_RELATIONS),
                        'name': f"가족{random.randint(1, 10)}",
                        'birth_date': generate_birth_date(1, 80),
                        'occupation': random.choice(['회사원', '공무원', '자영업', '학생', '주부', None]),
                        'contact': generate_phone() if random.random() > 0.5 else None,
                        'is_cohabitant': random.choice([True, False]),
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    family_id += 1
            
            # HrProject (인사이력 프로젝트) - 1-3개
            for _ in range(random.randint(1, 3)):
                start_date = generate_hire_date(max_years=2)
                hr_projects.append({
                    'id': hr_project_id,
                    'employee_id': emp_id,
                    'project_name': random.choice(PROJECT_NAMES),
                    'start_date': start_date,
                    'end_date': None if random.random() > 0.7 else generate_hire_date(max_years=1),
                    'duration': f"{random.randint(3, 24)}개월",
                    'role': random.choice(PROJECT_ROLES),
                    'duty': f"{random.choice(['개발', '기획', '관리', '분석'])} 업무",
                    'client': f"{random.choice(['삼성', 'LG', 'SK', 'KT', '현대', '롯데'])}그룹",
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                hr_project_id += 1
            
            # ProjectParticipation (경력 프로젝트 참여) - 0-2개
            for _ in range(random.randint(0, 2)):
                project_participations.append({
                    'id': project_participation_id,
                    'employee_id': emp_id,
                    'project_name': f"이전회사 {random.choice(PROJECT_NAMES)}",
                    'start_date': generate_hire_date(max_years=5),
                    'end_date': generate_hire_date(max_years=3),
                    'role': random.choice(PROJECT_ROLES),
                    'duty': f"{random.choice(['개발', '기획', '관리'])} 담당",
                    'client': '이전고객사',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                project_participation_id += 1
            
            user_id += 1
            emp_id += 1
            profile_id += 1
    
    return (users, employees, profiles, educations, careers, certificates, languages, military_services,
            contracts, promotions, evaluations, trainings, attendances, salaries, salary_histories,
            salary_payments, benefits, insurances, assets, awards, family_members, hr_projects, project_participations)


def create_personal_accounts(resident_numbers):
    """개인 계정 데이터 생성"""
    users = []
    profiles = []
    personal_profiles = []
    educations = []
    careers = []
    certificates = []
    languages = []
    
    user_id = 100  # 직원과 구분하기 위해 100부터 시작
    profile_id = 100
    personal_profile_id = 1
    edu_id = 1000
    career_id = 1000
    cert_id = 1000
    lang_id = 1000
    
    rrn_idx = 60  # 직원 이후부터 사용
    
    # 남성 개인 계정 10명
    for male_idx in range(10):
        name = MALE_NAMES[30 + male_idx] if 30 + male_idx < len(MALE_NAMES) else f"남성{male_idx + 1}"
        gender = '남'
        birth_date = generate_birth_date()
        
        # 주민번호
        if rrn_idx < len(resident_numbers):
            rrn = resident_numbers[rrn_idx]
            rrn_idx += 1
        else:
            rrn = f"{birth_date.replace('-', '')[2:]}-1{random.randint(100000, 999999)}"
        
        # User 생성
        username = f"personal_{male_idx + 1}"
        users.append({
            'id': user_id,
            'username': username,
            'email': generate_email(name),
            'password_hash': 'pbkdf2:sha256:600000$dummy$dummy',
            'role': 'user',
            'account_type': 'personal',
            'company_id': None,
            'employee_id': None,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # Profile 생성
        profiles.append({
            'id': profile_id,
            'name': name,
            'gender': gender,
            'birth_date': birth_date,
            'resident_number': rrn,
            'mobile_phone': generate_phone(),
            'email': generate_email(name),
            'address': generate_address(),
            'postal_code': generate_postal_code(),
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # PersonalProfile 생성
        personal_profiles.append({
            'id': personal_profile_id,
            'user_id': user_id,
            'profile_id': profile_id,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # Education (1-2개)
        for _ in range(random.randint(1, 2)):
            educations.append({
                'id': edu_id,
                'employee_id': None,
                'profile_id': profile_id,
                'level': random.choice(EDUCATION_LEVELS),
                'school_name': random.choice(UNIVERSITIES),
                'major': random.choice(MAJORS),
                'graduation_date': f"{random.randint(2010, 2020)}-{random.randint(1, 12):02d}",
                'status': '졸업',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            edu_id += 1
        
        # Career (1-3개)
        for _ in range(random.randint(1, 3)):
            start_date = generate_hire_date(max_years=10)
            end_date = generate_hire_date(min_years=0, max_years=5)
            careers.append({
                'id': career_id,
                'employee_id': None,
                'profile_id': profile_id,
                'company_name': random.choice(CAREER_COMPANIES),
                'position': random.choice(POSITIONS),
                'start_date': start_date,
                'end_date': end_date,
                'description': '담당 업무 수행',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            career_id += 1
        
        # Certificate (0-2개)
        for _ in range(random.randint(0, 2)):
            cert_date = generate_hire_date(max_years=5)
            certificates.append({
                'id': cert_id,
                'employee_id': None,
                'profile_id': profile_id,
                'name': random.choice(CERTIFICATE_NAMES),
                'issuing_organization': '한국산업인력공단',
                'issue_date': cert_date,
                'expiry_date': None,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            cert_id += 1
        
        # Language (0-2개)
        for _ in range(random.randint(0, 2)):
            languages.append({
                'id': lang_id,
                'employee_id': None,
                'profile_id': profile_id,
                'language': random.choice(LANGUAGES),
                'level': random.choice(LANGUAGE_LEVELS),
                'test_name': 'TOEIC',
                'test_score': random.randint(600, 990) if random.choice(LANGUAGES) == '영어' else None,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            lang_id += 1
        
        user_id += 1
        profile_id += 1
        personal_profile_id += 1
    
    # 여성 개인 계정 10명
    for female_idx in range(10):
        name = FEMALE_NAMES[30 + female_idx] if 30 + female_idx < len(FEMALE_NAMES) else f"여성{female_idx + 1}"
        gender = '여'
        birth_date = generate_birth_date()
        
        # 주민번호
        if rrn_idx < len(resident_numbers):
            rrn = resident_numbers[rrn_idx]
            rrn_idx += 1
        else:
            rrn = f"{birth_date.replace('-', '')[2:]}-2{random.randint(100000, 999999)}"
        
        # User 생성
        username = f"personal_{female_idx + 11}"
        users.append({
            'id': user_id,
            'username': username,
            'email': generate_email(name),
            'password_hash': 'pbkdf2:sha256:600000$dummy$dummy',
            'role': 'user',
            'account_type': 'personal',
            'company_id': None,
            'employee_id': None,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # Profile 생성
        profiles.append({
            'id': profile_id,
            'name': name,
            'gender': gender,
            'birth_date': birth_date,
            'resident_number': rrn,
            'mobile_phone': generate_phone(),
            'email': generate_email(name),
            'address': generate_address(),
            'postal_code': generate_postal_code(),
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # PersonalProfile 생성
        personal_profiles.append({
            'id': personal_profile_id,
            'user_id': user_id,
            'profile_id': profile_id,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # Education (1-2개)
        for _ in range(random.randint(1, 2)):
            educations.append({
                'id': edu_id,
                'employee_id': None,
                'profile_id': profile_id,
                'level': random.choice(EDUCATION_LEVELS),
                'school_name': random.choice(UNIVERSITIES),
                'major': random.choice(MAJORS),
                'graduation_date': f"{random.randint(2010, 2020)}-{random.randint(1, 12):02d}",
                'status': '졸업',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            edu_id += 1
        
        # Career (1-3개)
        for _ in range(random.randint(1, 3)):
            start_date = generate_hire_date(max_years=10)
            end_date = generate_hire_date(min_years=0, max_years=5)
            careers.append({
                'id': career_id,
                'employee_id': None,
                'profile_id': profile_id,
                'company_name': random.choice(CAREER_COMPANIES),
                'position': random.choice(POSITIONS),
                'start_date': start_date,
                'end_date': end_date,
                'description': '담당 업무 수행',
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            career_id += 1
        
        # Certificate (0-2개)
        for _ in range(random.randint(0, 2)):
            cert_date = generate_hire_date(max_years=5)
            certificates.append({
                'id': cert_id,
                'employee_id': None,
                'profile_id': profile_id,
                'name': random.choice(CERTIFICATE_NAMES),
                'issuing_organization': '한국산업인력공단',
                'issue_date': cert_date,
                'expiry_date': None,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            cert_id += 1
        
        # Language (0-2개)
        for _ in range(random.randint(0, 2)):
            languages.append({
                'id': lang_id,
                'employee_id': None,
                'profile_id': profile_id,
                'language': random.choice(LANGUAGES),
                'level': random.choice(LANGUAGE_LEVELS),
                'test_name': 'TOEIC',
                'test_score': random.randint(600, 990) if random.choice(LANGUAGES) == '영어' else None,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            lang_id += 1
        
        user_id += 1
        profile_id += 1
        personal_profile_id += 1
    
    return users, profiles, personal_profiles, educations, careers, certificates, languages


def create_corporate_users(companies):
    """법인 계정 생성"""
    users = []
    user_id = 200  # 법인 계정은 200부터 시작
    
    for company in companies:
        username = f"corp_{company['id']}"
        users.append({
            'id': user_id,
            'username': username,
            'email': f"admin@{company['name'].lower().replace('주식회사 ', '').replace(' ', '')}.com",
            'password_hash': 'pbkdf2:sha256:600000$dummy$dummy',
            'role': 'manager',
            'account_type': 'corporate',
            'company_id': company['id'],
            'employee_id': None,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        user_id += 1
    
    return users


# ============================================================
# Excel 생성 함수
# ============================================================

def create_excel_file(data_dict, output_path):
    """Excel 파일 생성"""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for sheet_name, data_list in data_dict.items():
        if not data_list:
            continue
        
        ws = wb.create_sheet(title=sheet_name)
        
        # 헤더 작성
        headers = list(data_list[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_idx, row_data in enumerate(data_list, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                # None 값을 빈 문자열로 변환
                if value is None:
                    value = ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 컬럼 너비 자동 조정
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    wb.save(output_path)
    print(f"\n[완료] Excel 파일 생성: {output_path}")


# ============================================================
# 메인 함수
# ============================================================

def main():
    print("=" * 70)
    print("Excel 테스트 DB 생성 스크립트")
    print("=" * 70)
    
    # 주민번호 로드
    print("\n[1단계] 주민번호 로드")
    resident_numbers = load_resident_numbers()
    print(f"  로드된 주민번호: {len(resident_numbers)}개")
    
    # 데이터 생성
    print("\n[2단계] 데이터 생성")
    
    print("  - 법인 데이터 생성 중...")
    companies = create_companies()
    print(f"    생성된 법인: {len(companies)}개")
    
    print("  - 조직 데이터 생성 중...")
    organizations = create_organizations(companies)
    print(f"    생성된 조직: {len(organizations)}개")
    
    print("  - 법인 계정 생성 중...")
    corporate_users = create_corporate_users(companies)
    print(f"    생성된 법인 계정: {len(corporate_users)}개")
    
    print("  - 직원 및 사용자 데이터 생성 중...")
    (emp_users, employees, emp_profiles, educations, careers, certificates, languages, military_services,
     contracts, promotions, evaluations, trainings, attendances, salaries, salary_histories,
     salary_payments, benefits, insurances, assets, awards, family_members, hr_projects, project_participations) = \
        create_users_and_employees(companies, organizations, resident_numbers)
    print(f"    생성된 직원 사용자: {len(emp_users)}개")
    print(f"    생성된 직원: {len(employees)}개")
    print(f"    생성된 직원 프로필: {len(emp_profiles)}개")
    print(f"    생성된 학력: {len(educations)}개")
    print(f"    생성된 경력: {len(careers)}개")
    print(f"    생성된 자격증: {len(certificates)}개")
    print(f"    생성된 어학: {len(languages)}개")
    print(f"    생성된 병역: {len(military_services)}개")
    print(f"    생성된 계약: {len(contracts)}개")
    print(f"    생성된 발령: {len(promotions)}개")
    print(f"    생성된 평가: {len(evaluations)}개")
    print(f"    생성된 교육: {len(trainings)}개")
    print(f"    생성된 근태: {len(attendances)}개")
    print(f"    생성된 급여: {len(salaries)}개")
    print(f"    생성된 급여이력: {len(salary_histories)}개")
    print(f"    생성된 급여지급: {len(salary_payments)}개")
    print(f"    생성된 복리후생: {len(benefits)}개")
    print(f"    생성된 보험: {len(insurances)}개")
    print(f"    생성된 자산: {len(assets)}개")
    print(f"    생성된 수상: {len(awards)}개")
    print(f"    생성된 가족: {len(family_members)}개")
    print(f"    생성된 프로젝트: {len(hr_projects)}개")
    print(f"    생성된 프로젝트참여: {len(project_participations)}개")
    
    print("  - 개인 계정 데이터 생성 중...")
    personal_users, personal_profiles, personal_profile_records, personal_educations, personal_careers, personal_certificates, personal_languages = \
        create_personal_accounts(resident_numbers)
    print(f"    생성된 개인 사용자: {len(personal_users)}개")
    print(f"    생성된 개인 프로필: {len(personal_profiles)}개")
    print(f"    생성된 PersonalProfile: {len(personal_profile_records)}개")
    print(f"    생성된 개인 학력: {len(personal_educations)}개")
    print(f"    생성된 개인 경력: {len(personal_careers)}개")
    print(f"    생성된 개인 자격증: {len(personal_certificates)}개")
    print(f"    생성된 개인 어학: {len(personal_languages)}개")
    
    # 모든 사용자 합치기
    all_users = corporate_users + emp_users + personal_users
    all_profiles = emp_profiles + personal_profiles
    all_educations = educations + personal_educations
    all_careers = careers + personal_careers
    all_certificates = certificates + personal_certificates
    all_languages = languages + personal_languages
    
    # Excel 파일 생성
    print("\n[3단계] Excel 파일 생성")
    data_dict = {
        'Companies': companies,
        'Organizations': organizations,
        'Users': all_users,
        'Profiles': all_profiles,
        'PersonalProfiles': personal_profile_records,
        'Employees': employees,
        'Educations': all_educations,
        'Careers': all_careers,
        'Certificates': all_certificates,
        'Languages': all_languages,
        'MilitaryServices': military_services,
        # 인사관리 테이블
        'Contracts': contracts,
        'Promotions': promotions,
        'Evaluations': evaluations,
        'Trainings': trainings,
        'Attendances': attendances,
        'Salaries': salaries,
        'SalaryHistories': salary_histories,
        'SalaryPayments': salary_payments,
        'Benefits': benefits,
        'Insurances': insurances,
        'Assets': assets,
        'Awards': awards,
        'FamilyMembers': family_members,
        'HrProjects': hr_projects,
        'ProjectParticipations': project_participations,
    }
    
    create_excel_file(data_dict, OUTPUT_EXCEL_PATH)
    
    # 요약 출력
    print("\n" + "=" * 70)
    print("생성 완료 요약")
    print("=" * 70)
    print(f"법인: {len(companies)}개")
    print(f"조직: {len(organizations)}개")
    print(f"사용자: {len(all_users)}개 (법인 {len(corporate_users)} + 직원 {len(emp_users)} + 개인 {len(personal_users)})")
    print(f"프로필: {len(all_profiles)}개")
    print(f"직원: {len(employees)}개")
    print(f"학력: {len(all_educations)}개")
    print(f"경력: {len(all_careers)}개")
    print(f"자격증: {len(all_certificates)}개")
    print(f"어학: {len(all_languages)}개")
    print(f"병역: {len(military_services)}개")
    print(f"\n[인사관리 데이터]")
    print(f"계약: {len(contracts)}개")
    print(f"발령: {len(promotions)}개")
    print(f"평가: {len(evaluations)}개")
    print(f"교육: {len(trainings)}개")
    print(f"근태: {len(attendances)}개")
    print(f"급여: {len(salaries)}개")
    print(f"급여이력: {len(salary_histories)}개")
    print(f"급여지급: {len(salary_payments)}개")
    print(f"복리후생: {len(benefits)}개")
    print(f"보험: {len(insurances)}개")
    print(f"자산: {len(assets)}개")
    print(f"수상: {len(awards)}개")
    print(f"가족: {len(family_members)}개")
    print(f"프로젝트: {len(hr_projects)}개")
    print(f"프로젝트참여: {len(project_participations)}개")
    print(f"\nExcel 파일: {OUTPUT_EXCEL_PATH}")
    print("=" * 70)


if __name__ == '__main__':
    main()
